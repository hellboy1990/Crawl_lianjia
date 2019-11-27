
###---@LJ出品，转载请注明---###

'''
此程序为爬取链家网上各城市的小区及其房屋信息，包括图片及文字部分
'''

import pprint
import requests
from requests import RequestException
from bs4 import BeautifulSoup
import pandas as pd
import re
from datetime import datetime
import os
import json
from headers import get_headers
import random
import math
import openpyxl
import time
tim = random.randint(3,5)


def req(url):#定义请求函数
    headers = {'User-Agent': get_headers()}
    webdata = requests.get(url, headers=headers, timeout=30)
    # print(webdata)
    return webdata


def get_district(url):#获取各城区链接
    global dis_links,districts
    districts=[]
    dis_links=[]
    webdata=req(url)
    if webdata.status_code == 200:
        soups = BeautifulSoup(webdata.text, 'lxml')
        dists = soups.select('dl dd div a')
        for dist, link in [[i.get_text(), domain + i.get('href')] for i in dists[:-2]]:
            print(dist, link)
            districts.append(dist)
            dis_links.append(link)
    else:
        pass


def get_areas(area_infos):#获取城区的各个区域
    wb=openpyxl.Workbook()
    wbs=wb.active
    wbs.append(['城区','城区链接','区域','区域链接'])
    #for i in range(0,1):
    for i in range(0,len(dis_links)):
        webdata=req(dis_links[i])
        if webdata.status_code == 200:
            soups=BeautifulSoup(webdata.text,'lxml')
            areas=soups.select('dl dd div div:nth-of-type(2) a')
            #print(areas)
            for area in areas:
                area_name=area.get_text()
                area_link=domain+area.get('href')
                area_i=[districts[i],dis_links[i],area_name,area_link]
                print(area_i)
                wbs.append(area_i)
        else:
            pass
        time.sleep(tim*1.2)    #print(areas_list,areas_links,dists_list)
    wb.save(area_infos)


def get_xiaoqu(area_infos,file_xq):#获取各个区域的小区
    #读取城区、区域等信息
    dfs=pd.read_excel(area_infos,sheet_name=0)
    dists_list=list(dfs["城区"])
    #districts=list(dfs["城区链接"])
    areas_list=list(dfs["区域"])
    areas_links=list(dfs["区域链接"])
    #新建表格
    wb=openpyxl.Workbook()
    wbs=wb.active
    wbs.append(['城区','区域','区域链接','小区名称','小区链接'])
    for i in range(33,34):
    #for i in range(34, len(areas_links)):
    #for i in range(0,len(areas_links)):#构造区域循环
        print(areas_list[i],areas_links[i])
        webdata=req(areas_links[i])
        if webdata.status_code == 200:
            try:
                soups = BeautifulSoup(webdata.text, 'lxml')
                xq_pages = soups.select('div.page-box.house-lst-page-box')
                #print(len(xq_pages),type(json.loads(xq_pages[0].get('page-data'))['totalPage']))#页数是整型
                '''判断共有几页，如果有一页则直接爬取链接，否则先确定页数再爬取链接'''
                if len(xq_pages) == 0:  # 如果只有一页
                    print(areas_list[i]+'有一页！')
                    xqs= soups.select(' li.clear.LOGCLICKDATA div.info div.title a')
                    for xq in xqs:
                        xq_title = xq.get_text()
                        xq_link = xq.get('href')
                        xq_i = [dists_list[i], areas_list[i], xqs, xq_title, xq_link]
                        print(xq_i)
                        wbs.append(xq_i)
                else:  # 大于一页
                    xq_pages = json.loads(xq_pages[0].get('page-data'))['totalPage']
                    print(areas_list[i]+'有'+str(xq_pages)+'页！')
                    for page in range(1, int(xq_pages)+1):# 构造翻页循环
                        xq_url =areas_links[i]+'pg'+str(page)
                        #print(xq_url)
                        res=req(xq_url)
                        try:
                            res_soups=BeautifulSoup(res.text,'lxml')
                            xqs=res_soups.select('li.clear div.info div.title a')#定位小区信息列表
                            for xq in xqs:
                                xq_title=xq.get_text()
                                xq_link=xq.get('href')
                                xq_i=[dists_list[i],areas_list[i],xq_url,xq_title,xq_link]
                                print(xq_i)
                                wbs.append(xq_i)
                        except:
                            continue
            except:
                wb.save(file_xq)
                continue
        else:
            pass
        time.sleep(tim*1.5)
    wb.save(file_xq)
    print('已爬取城市区域内所有小区链接！')


def get_zs_cj(file_xq,file_xq_info):#获取小区已成效和在售的全部房屋链接
    #读取小区信息
    xq_infos= pd.read_excel(file_xq)
    xq_links=list(xq_infos['小区链接'])
    xq_titles=list(xq_infos['小区名称'])
    dists_list=list(xq_infos['城区'])
    areas_list=list(xq_infos['区域'])
    #建立表格
    wb=openpyxl.Workbook()
    wbs=wb.active
    wbs.append(['城区','区域','小区名称','小区均价','参考月份','建筑年代','建筑类型','物业费用','物业公司',
                '开发商','楼栋总数','房屋总数','在售链接','成交链接'])#,'图片列表'
    #for i in range(0,1):#构造小区循环
    for i in range(4284, len(xq_links)):
        print(xq_titles[i])
        webdata=req(xq_links[i])
        if webdata.status_code==200:
            try:
                soups=BeautifulSoup(webdata.text,'lxml')
                link_zaishou=soups.select('div.goodSell a.fr')[0].get('href')#在售链接
                #print(link_zaishou)
                link_chengjiao = soups.select('div.frameDeal a.btn-large')[0].get('href')#成交链接
                #print(link_chengjiao)
                unitprice = soups.select('div.xiaoquPrice span')[0].text.strip()#小区均价
                #print(unitprice)
                pricetime = soups.select('div.xiaoquPrice span')[1].text.strip()#参考月份
                #print(pricetime)
                jznd = soups.select('div.xiaoquInfoItem span.xiaoquInfoContent')[0].text.strip()#建筑年代
                jzlx = soups.select('div.xiaoquInfoItem span.xiaoquInfoContent')[1].text.strip()#建筑类型
                wyfy = soups.select('div.xiaoquInfoItem span.xiaoquInfoContent')[2].text.strip()#物业费用
                wygs = soups.select('div.xiaoquInfoItem span.xiaoquInfoContent')[3].text.strip()#物业公司
                kfs = soups.select('div.xiaoquInfoItem span.xiaoquInfoContent')[4].text.strip()#开发商
                ldzs = soups.select('div.xiaoquInfoItem span.xiaoquInfoContent')[5].text.strip()#楼栋总数
                fwzs = soups.select('div.xiaoquInfoItem span.xiaoquInfoContent')[6].text.strip()#房屋总数
                #print(jznd,fwzs)
                # # 图片链接(图片链接目前先不写入,因为字符数太多)
                # img_link = soups.select('ol#overviewThumbnail li')
                #img_links=[link.get('data-src') for link in img_link]#图片列表
                #print(img_links)
                xq_infos_i = [dists_list[i],areas_list[i],xq_titles[i],unitprice,pricetime,
                            jznd,jzlx,wyfy,wygs,kfs,ldzs,fwzs,link_zaishou,link_chengjiao]#,img_links
                wbs.append(xq_infos_i)
                #print(xq_infos_i)
            except:
                wb.save(file_xq_info)
                print('Error!')
                continue
        else:
            pass
        time.sleep(tim*1.2)
    wb.save(file_xq_info)
    print('已爬取让所有小区信息及图片链接！')


# def get_zs_link(xq_infos_i,file_zs):#获取小区在售房屋的所有链接
#     zs_links=[]
#     for i in range(0, len(links_zaishou)):  # 构造小区在售链接循环
#         print(links_zaishou[i])
#         webdata = req(links_zaishou[i])
#         if webdata.status_code == 200:
#             soups = BeautifulSoup(webdata.text, 'lxml')
#             zs_pages=soups.select('div.page-box.house-lst-page-box')
#             '''判断共有几页，如果有一页则直接爬取链接，否则先确定页数再爬取链接'''
#             if len(zs_pages)==0:#如果只有一页
#                 print(xqnames[i] + '有一页！')
#                 zs_urls=soups.select('ul.sellListContent li.clear.LOGCLICKDATA div.title a')
#                 for zs_url in zs_urls:
#                     zs_links.append(zs_url.get('href'))
#             else:#大于一页
#                 zs_pages = json.loads(zs_pages[0].get('page-data'))['totalPage']
#                 print(xqnames[i]+'有'+str(zs_pages)+'页！')
#                 for page in range(1,zs_pages+1):#构造翻页循环
#                     zs_page = 'ershoufang/pg' + str(page)
#                     zs_link = links_zaishou[i].replace('ershoufang/', zs_page)
#                     zs_res = req(zs_link)
#                     try:
#                         cj_soups = BeautifulSoup(zs_res.text, 'lxml')
#                         zs_urls = cj_soups.select('ul.sellListContent li.clear.LOGCLICKDATA div.title a')
#                         for zs_url in zs_urls:  # 构造成效房屋循环
#                             zs_links.append(zs_url.get('href'))
#                     except:
#                         continue
#                     time.sleep(tim)
#         else:
#             pass
#         time.sleep(tim)
#     #print(zs_links,len(zs_links))
#     df=pd.DataFrame({'在售链接':zs_links})
#     print(df.head())
#     df.to_excel(os.getcwd()+file_zs)
#     print('已爬取所有小区在售住房链接！')
#
#
def get_cj_link(file_xq_info,file_cj):#获取小区成交房屋的所有链接
    links_chengjiao = pd.read_excel(file_xq_info)['成交链接']
    dists_list = pd.read_excel(file_xq_info)['城区']
    areas_list = pd.read_excel(file_xq_info)['区域']
    xqnames=pd.read_excel(file_xq_info)['小区名称']
    # 建立表格
    wb = openpyxl.Workbook()
    wbs = wb.active
    wbs.append(['城区', '区域', '小区名称', '成交链接'])
    #for i in range(0,10):
    for i in range(4731,len(links_chengjiao)):#构造成交页循环
        #print(links_chengjiao[i])
        webdata=req(links_chengjiao[i])
        if webdata.status_code==200:
            soups=BeautifulSoup(webdata.text,'lxml')
            cj_pages=soups.select('div.page-box.house-lst-page-box')
            #print(cj_pages)
            try:
                '''判断共有几页，如果有一页则直接爬取链接，否则先确定页数再爬取链接'''
                if len(cj_pages) == 0:  # 如果只有一页
                    print(xqnames[i] + '有一页！')
                    cj_urls = soups.select('li div.info div.title a')
                    for cj_url in cj_urls:
                        cj_link=cj_url.get('href')
                        cj_i=[dists_list[i],areas_list[i],xqnames[i],cj_link]
                        wbs.append(cj_i)
                else:
                    cj_pages = json.loads(cj_pages[0].get('page-data'))['totalPage']
                    print(xqnames[i] + '有' + str(cj_pages) + '页！')
                    for page in range(1,cj_pages+1):#构造翻页循环
                        cj_page='chengjiao/pg'+str(page)
                        cj_link=links_chengjiao[i].replace('chengjiao/',cj_page)
                        try:
                            cj_res=req(cj_link)
                            cj_soups=BeautifulSoup(cj_res.text,'lxml')
                            cj_urls=cj_soups.select('li div.info div.title a')
                            for cj_url in cj_urls:#构造成交房屋循环
                                cj_link = cj_url.get('href')
                                cj_i = [dists_list[i], areas_list[i], xqnames[i], cj_link]
                                wbs.append(cj_i)
                        except:
                            wb.save(file_cj)
            except:
                wb.save(file_cj)
                print('Error!')
                continue
            time.sleep(tim)
        else:
            pass
        time.sleep(tim*1.5)
    wb.save(file_cj)
    print('已爬取所有小区成交住房链接！')


def get_zs_house():
    info_jznds = []  #建筑年代
    xqnms = []  # 小区名称
    totals = []  # 总价
    unitprices = []  # 单价
    info_fwhxs = []  # 房屋户型
    info_fwlcs = []  # 所在楼层
    info_fwmjs = []  # 建筑面积
    info_hxjgs = []  # 户型结构
    info_tnmjs = []  # 套内面积
    info_leixs = []  # 建筑类型
    info_chaos = []  # 房屋朝向
    info_jzjgs = []  # 建筑结构
    info_zxius = []  # 装修情况
    info_thbls = []  # 梯户比例
    info_gnfss = []  # 供暖方式
    info_pbdts = []  # 配备电梯
    info_cqnxs = []  # 产权年限
    # tr_gpsjs = []  # 挂牌时间
    # tr_jyqss = []  # 交易权属
    # tr_scjys = []  # 上次交易
    # tr_fwyts = []  # 房屋用途
    # tr_fwnxs = []  # 房屋年限
    # tr_cqsss = []  # 产权所属
    # tr_dyxxs = []  # 抵押信息
    # tr_fbbjs = []  # 房本备件
    house_link=pd.read_excel(file_zs)['在售链接']
    for i in range(0,len(house_link)):
        print(house_link[i])
        webdata=req(house_link[i])
        if webdata.status_code==200:
            soups = BeautifulSoup(webdata.text, 'lxml')
            try:
                total=soups.select('div.price span.total')[0].text.strip()#总价
                totals.append(total)
                #print(totals)
                unitprice=soups.select('div.price div.unitPrice span')[0].text.strip()#单价
                unitprices.append(unitprice)
                info_jznd = soups.select('div.area div.subInfo')[0].text.strip()#建筑年代
                info_jznds.append(info_jznd)
                xqnm= soups.select('div.aroundInfo div.communityName')#小区名称
                xqnms.append(xqnm)
                #基本信息
                info_fwhx = soups.select('div.base div.content ul li span')[0].next_sibling#户型
                info_fwhxs.append(info_fwhx)
                # print(info_fwhxs)
                info_fwlc = soups.select('div.base div.content ul li span')[1].next_sibling#楼层
                info_fwlcs.append(info_fwlc)
                info_fwmj = soups.select('div.base div.content ul li span')[2].next_sibling
                info_fwmjs.append(info_fwmj)
                info_hxjg = soups.select('div.base div.content ul li span')[3].next_sibling
                info_hxjgs.append(info_hxjg)
                info_tnmj = soups.select('div.base div.content ul li span')[4].next_sibling
                info_tnmjs.append(info_tnmj)
                info_leix = soups.select('div.base div.content ul li span')[5].next_sibling
                info_leixs.append(info_leix)
                info_chao = soups.select('div.base div.content ul li span')[6].next_sibling
                info_chaos.append(info_chao)
                info_jzjg = soups.select('div.base div.content ul li span')[7].next_sibling
                info_jzjgs.append(info_jzjg)
                info_zxiu = soups.select('div.base div.content ul li span')[8].next_sibling
                info_zxius.append(info_zxiu)
                info_thbl = soups.select('div.base div.content ul li span')[9].next_sibling
                info_thbls.append(info_thbl)
                info_gnfs = soups.select('div.base div.content ul li span')[10].next_sibling
                info_gnfss.append(info_gnfs)
                info_pbdt = soups.select('div.base div.content ul li span')[11].next_sibling
                info_pbdts.append(info_pbdt)
                info_cqnx = soups.select('div.base div.content ul li span')[12].next_sibling
                info_cqnxs.append(info_cqnx)
            except:
                continue
        time.sleep(tim*1.2)
    print(len(xqnms),len(info_chaos))
    # 写入文件
    df_house = pd.DataFrame({'name': xqnms, 'tot': totals, 'unit': unitprices, 'jznt': info_jznds, 'fwhx': info_fwhxs,
                       'fwlc': info_fwlcs, 'fwmj': info_fwmjs, 'hxjg': info_hxjgs, 'tnmj': info_tnmjs,
                       'leixin': info_leixs, 'chaoxiang': info_chaos, 'jzjg': info_jzjgs, 'zhuangxiu': info_zxius,
                       'tnbl': info_thbls, 'gnfs': info_gnfss, 'pbdt': info_pbdts, 'cqnx': info_cqnxs,})
    print(df_house.head())
    df_house.to_csv(os.getcwd()+file_zs_house, index=False, sep=',', encoding='utf-8')
    print('已爬取所有在售房屋信息！')


def get_cj_house(file_cj,file_cj_house):
    # 建立表格
    wb = openpyxl.Workbook()
    wbs = wb.active
    wbs.append(['城区','区域','小区名称','成效时间','交易总价', '挂牌价格', '成交周期', '调价', '带看', '关注', '浏览',
                '房屋户型', '所在楼层',
                '建筑面积', '户型结构', '套内面积', '建筑类型', '房屋朝向','建成年代','装修情况','建筑结构',
                '供暖方式','梯户比例','产权年限'])
    dists_list = pd.read_excel(file_cj)['城区']
    areas_list = pd.read_excel(file_cj)['区域']
    xqnames = pd.read_excel(file_cj)['小区名称']
    house_link = pd.read_excel(file_cj)['成交链接']#div.house-title span
    #for i in range(0,1):
    for i in range(35767, len(house_link)):
        try:
            print(i)
            #print(house_link[i])
            webdata = req(house_link[i])
            if webdata.status_code == 200:
                soups = BeautifulSoup(webdata.text, 'lxml')
                dealtime=soups.select('div.house-title span')[0].text.strip()#成效时间
                dealprice= soups.select('div.price span.dealTotalPrice')[0].text.strip()  # 交易总价
                guapai=soups.select('div.info.fr div.msg span')[0].text.strip()
                chengjiao=soups.select('div.info.fr div.msg span')[1].text.strip()
                tiaojia=soups.select('div.info.fr div.msg span')[2].text.strip()
                daikan= soups.select('div.info.fr div.msg span')[3].text.strip()
                guanzhu = soups.select('div.info.fr div.msg span')[4].text.strip()
                liulan= soups.select('div.info.fr div.msg span')[5].text.strip()
                # 基本信息
                info_fwhx = soups.select('div.base div.content ul li span')[0].next_sibling  # 户型
                info_fwlc = soups.select('div.base div.content ul li span')[1].next_sibling  # 楼层
                info_fwmj = soups.select('div.base div.content ul li span')[2].next_sibling
                info_hxjg = soups.select('div.base div.content ul li span')[3].next_sibling
                info_tnmj = soups.select('div.base div.content ul li span')[4].next_sibling
                info_leix = soups.select('div.base div.content ul li span')[5].next_sibling
                info_chao = soups.select('div.base div.content ul li span')[6].next_sibling
                info_jzjg = soups.select('div.base div.content ul li span')[7].next_sibling
                info_zxiu = soups.select('div.base div.content ul li span')[8].next_sibling
                info_thbl = soups.select('div.base div.content ul li span')[9].next_sibling
                info_gnfs = soups.select('div.base div.content ul li span')[10].next_sibling
                info_pbdt = soups.select('div.base div.content ul li span')[11].next_sibling
                info_cqnx = soups.select('div.base div.content ul li span')[12].next_sibling
                house_i=[dists_list[i],areas_list[i],xqnames[i],dealtime,dealprice,guapai,chengjiao,tiaojia,daikan,guanzhu,
                         liulan,info_fwhx,info_fwlc,info_fwmj,
                         info_hxjg,info_tnmj,info_leix,info_chao,info_jzjg,info_zxiu,info_thbl,info_gnfs,info_pbdt,
                         info_cqnx]
                wbs.append(house_i)
                #print(house_i)
        except:
            print('Error!')
            wb.save(file_cj_house)
            continue
        time.sleep(tim * 1.5)
    wb.save(file_cj_house)
    print('已爬取所有在售房屋信息！')


if __name__=="__main__":
    time_start = datetime.now()
    print('开始时间:' + str(time_start))
    cityname=input("请输入待爬取取城市简写：")
    #事先生成一个表格
    file_xq = os.getcwd() + '.\\%s_xq_infos.xlsx'%cityname
    file_zs = os.getcwd() + '.\\%s_zs_links.xlsx'%cityname
    file_cj = os.getcwd() + '.\\%s_cj_links.xlsx'%cityname
    file_area=os.getcwd() + '.\\%s_areas_infos.xlsx'%cityname
    file_xq_info = os.getcwd() + '.\\%s_xq_housrinfos.xlsx'%cityname
    file_zs_house = '.\\%s_zs_house.xlsx'%cityname
    file_cj_house = '.\\%s_cj_house.xlsx'%cityname
    #print(file_xq)
    # 第一步爬取所有城区的区域内的所有小区并保存到列表
    step1=input("是否开始爬取：")
    if int(step1)==0:
        domain='https://%s.lianjia.com'%cityname
        url = 'https://%s.lianjia.com/xiaoqu/'%cityname
        print(url)
        #get_district(url)
        # print('已获取各区链接！')
        # get_areas(file_area)
        # print('已获取各城区的区域链接！')
        #get_xiaoqu(file_area,file_xq)
        #print('已获取各区域的小区链接！')
        #第二步爬取所有小区房屋的成交和在售列表，以及图片链接
        step2=input("是否继续爬取房屋链接：")
        if int(step2)==0:
            #get_zs_cj(file_xq,file_xq_info)
            #get_zs_link(file_cj,file_cj_house)
            #get_cj_link(file_xq_info,file_cj)
            #第三步是爬取所有房屋信息
            step3 = input("是否开始继续爬取房屋信息：")
            if int(step3) == 0:
                #get_zs_house()
                get_cj_house(file_cj,file_cj_house)
            else:
                pass
        else:
            pass
    else:
        pass
    time_end = datetime.now()
    print('结束时间:' + str(time_end))
    time_last = time_end - time_start
    print('用时' + str(time_last))